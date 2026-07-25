import json
import os
import re
import time

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form
from openpyxl import load_workbook

from app.db import query
from app.security import get_current_user

router = APIRouter()

UPLOAD_DIR = os.path.join(os.path.dirname(__file__), "..", "..", "uploads")
os.makedirs(UPLOAD_DIR, exist_ok=True)


@router.get("")
def list_reports(user=Depends(get_current_user)):
    try:
        return query(
            """SELECT id, title, period, month, year, file_name, sheet_name, created_at,
                      jsonb_array_length(data->'rows') as row_count
               FROM quarterly_reports
               ORDER BY created_at DESC"""
        )
    except Exception as err:
        raise HTTPException(status_code=500, detail=str(err))


@router.get("/{id}")
def get_report(id: int, user=Depends(get_current_user)):
    try:
        rows = query("SELECT * FROM quarterly_reports WHERE id=%s", [id])
        if not rows:
            raise HTTPException(status_code=404, detail="Report not found")
        return rows[0]
    except HTTPException:
        raise
    except Exception as err:
        raise HTTPException(status_code=500, detail=str(err))


@router.post("/import")
async def import_report(
    user=Depends(get_current_user),
    excelFile: UploadFile = File(...),
    period: str = Form(default=None),
):
    if not excelFile:
        raise HTTPException(status_code=400, detail="Excel file is required")

    dest_path = os.path.join(UPLOAD_DIR, f"{int(time.time() * 1000)}-{excelFile.filename}")
    contents = await excelFile.read()
    with open(dest_path, "wb") as f:
        f.write(contents)

    try:
        wb = load_workbook(dest_path, data_only=True)
        sheet_name = wb.sheetnames[0]
        ws = wb[sheet_name]

        rows_data = []
        for row in ws.iter_rows(values_only=True):
            rows_data.append(["" if v is None else v for v in row])

        title_row = rows_data[0] if rows_data else []
        title = str(title_row[0]).strip() if title_row and title_row[0] != "" else "Quarterly Report"
        month_year = re.search(r"([A-Za-z]+)\s+(\d{4})", title)

        report_period = period or (title.split("\t")[0] if title else None)
        month = month_year.group(1) if month_year else None
        year = int(month_year.group(2)) if month_year else None

        inserted = query(
            """INSERT INTO quarterly_reports (title, period, month, year, file_name, sheet_name, data)
               VALUES (%s,%s,%s,%s,%s,%s,%s)
               RETURNING *""",
            [title, report_period, month, year, excelFile.filename, sheet_name, json.dumps({"rows": rows_data}, default=str)],
        )

        return {"report": inserted[0], "message": f"Imported {len(rows_data)} rows from {sheet_name}"}
    except HTTPException:
        raise
    except Exception as err:
        print(f"Quarterly import failed: {err}")
        raise HTTPException(status_code=500, detail=str(err))
